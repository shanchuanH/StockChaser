"""Extract ticker universe from AI_Chain_Watchlist.xlsx → data/universe.json.

Run once after editing the xlsx; output is consumed by fetch_data.py.
"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "AI_Chain_Watchlist.xlsx"
OUT = ROOT / "data" / "universe.json"

wb = load_workbook(XLSX, data_only=True)
ws = wb["AI链全景清单"]

rows = []
for r in range(2, ws.max_row + 1):
    ticker = ws.cell(row=r, column=2).value
    if not ticker:
        continue
    rows.append({
        "ticker": ticker,
        "name": ws.cell(row=r, column=3).value,
        "layer": ws.cell(row=r, column=4).value,
        "subsector": ws.cell(row=r, column=5).value,
        "role": ws.cell(row=r, column=6).value,
        # convert ★★★★★ → 5
        "potential": (ws.cell(row=r, column=7).value or "").count("★"),
        "strategy": ws.cell(row=r, column=8).value,
        "logic": ws.cell(row=r, column=9).value,
        "catalyst": ws.cell(row=r, column=10).value,
        "cn_pair": ws.cell(row=r, column=11).value,
        "risk": ws.cell(row=r, column=12).value,
    })

OUT.parent.mkdir(exist_ok=True)
OUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(rows)} tickers → {OUT}")
